"""XLSX/XLS document loader using openpyxl and pandas."""
from typing import Dict, Any, Optional, List
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

# Check if openpyxl is available
OPENPYXL_AVAILABLE = False
OPENPYXL_UNAVAILABLE_REASON = None

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    OPENPYXL_UNAVAILABLE_REASON = f"openpyxl not installed: {str(e)}. Install with: pip install openpyxl"

# Check if pandas is available (for XLS support)
PANDAS_AVAILABLE = False
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pass


class XLSXLoader:
    """Load XLSX/XLS documents using openpyxl."""
    
    def __init__(self):
        """Initialize XLSX loader."""
        self.name = "xlsx"
    
    def extract_text(self, file_path: str) -> Dict[str, Any]:
        """
        Extract text from XLSX/XLS document.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with extracted content
        """
        path = Path(file_path)
        extension = path.suffix.lower()
        
        # Use pandas for XLS files
        if extension == '.xls':
            return self._extract_with_pandas(file_path)
        
        if not OPENPYXL_AVAILABLE:
            if PANDAS_AVAILABLE:
                return self._extract_with_pandas(file_path)
            return {
                "success": False,
                "loader": self.name,
                "error": OPENPYXL_UNAVAILABLE_REASON
            }
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            pages = []
            tables = []
            text_parts = []
            table_index = 0
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                
                # Get all data from sheet
                rows = list(sheet.iter_rows(values_only=True))
                
                if not rows:
                    continue
                
                # First row as headers
                headers = [str(cell) if cell is not None else "" for cell in rows[0]]
                data_rows = []
                
                for row in rows[1:]:
                    data_rows.append([str(cell) if cell is not None else "" for cell in row])
                
                # Create table
                tables.append({
                    "page_number": sheet_idx + 1,
                    "table_index": table_index,
                    "headers": headers,
                    "rows": data_rows,
                    "caption": sheet_name,
                    "row_count": len(data_rows),
                    "col_count": len(headers)
                })
                table_index += 1
                
                # Create text representation
                sheet_text = f"Sheet: {sheet_name}\n"
                sheet_text += "\t".join(headers) + "\n"
                for row in data_rows:
                    sheet_text += "\t".join(row) + "\n"
                
                text_parts.append(sheet_text)
                
                pages.append({
                    "page_number": sheet_idx + 1,
                    "text": sheet_text,
                    "char_count": len(sheet_text)
                })
            
            wb.close()
            
            full_text = "\n\n".join(text_parts)
            
            return {
                "success": True,
                "loader": self.name,
                "metadata": {
                    "format": extension.lstrip('.'),
                    "sheet_count": len(wb.sheetnames),
                    "sheet_names": wb.sheetnames
                },
                "pages": pages,
                "tables": tables,
                "full_text": full_text,
                "total_pages": len(pages),
                "total_chars": len(full_text)
            }
            
        except Exception as e:
            logger.error(f"XLSX extraction failed: {str(e)}", exc_info=True)
            # Try pandas as fallback
            if PANDAS_AVAILABLE:
                return self._extract_with_pandas(file_path)
            return {
                "success": False,
                "loader": self.name,
                "error": str(e)
            }
    
    def _extract_with_pandas(self, file_path: str) -> Dict[str, Any]:
        """Extract using pandas as fallback."""
        if not PANDAS_AVAILABLE:
            return {
                "success": False,
                "loader": self.name,
                "error": "pandas not available for XLS extraction"
            }
        
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            
            pages = []
            tables = []
            text_parts = []
            table_index = 0
            
            for sheet_idx, sheet_name in enumerate(excel_file.sheet_names):
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                # Create table
                tables.append({
                    "page_number": sheet_idx + 1,
                    "table_index": table_index,
                    "headers": list(df.columns),
                    "rows": df.fillna("").astype(str).values.tolist(),
                    "caption": sheet_name,
                    "row_count": len(df),
                    "col_count": len(df.columns)
                })
                table_index += 1
                
                # Create text representation
                sheet_text = f"Sheet: {sheet_name}\n{df.to_string(index=False)}"
                text_parts.append(sheet_text)
                
                pages.append({
                    "page_number": sheet_idx + 1,
                    "text": sheet_text,
                    "char_count": len(sheet_text)
                })
            
            full_text = "\n\n".join(text_parts)
            
            return {
                "success": True,
                "loader": self.name,
                "metadata": {
                    "format": Path(file_path).suffix.lstrip('.'),
                    "sheet_count": len(excel_file.sheet_names),
                    "sheet_names": excel_file.sheet_names
                },
                "pages": pages,
                "tables": tables,
                "full_text": full_text,
                "total_pages": len(pages),
                "total_chars": len(full_text)
            }
            
        except Exception as e:
            logger.error(f"Pandas Excel extraction failed: {str(e)}", exc_info=True)
            return {
                "success": False,
                "loader": self.name,
                "error": str(e)
            }
    
    def is_available(self) -> bool:
        """Check if loader is available."""
        return OPENPYXL_AVAILABLE or PANDAS_AVAILABLE
    
    def get_unavailable_reason(self) -> Optional[str]:
        """Get reason why loader is unavailable."""
        if OPENPYXL_AVAILABLE or PANDAS_AVAILABLE:
            return None
        return OPENPYXL_UNAVAILABLE_REASON


# Global instance
xlsx_loader = XLSXLoader()
